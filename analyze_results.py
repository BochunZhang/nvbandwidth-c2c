#!/usr/bin/env python3
"""
analyze_results.py
==================
Parses nvbandwidth JSON result files and writes Excel reports.

Directory layout expected:
  results/
    standard/{CATEGORY}/{testname}.json   <- single-stream tests
    custom/{CATEGORY}/{testname}.json     <- parallel-stream tests

Outputs:
  results/standard/{CATEGORY}/{CATEGORY}.xlsx  (per-category)
  results/custom/{CATEGORY}/{CATEGORY}.xlsx    (per-category)
  results/summary.xlsx                          (all categories, one sheet each)

JSON schema (nvbandwidth --json output):
  {
    "nvbandwidth": {
      "testcases": [
        {
          "name": "<key>",
          "status": "Passed" | "Waived" | "Error",
          "bandwidth_description": "<last description>",
          "bandwidth_matrix": [
            ["<float_or_NA>", ...],   <- row 0 (stream 0 / src GPU 0)
            ["<float_or_NA>", ...],   <- row 1 (stream 1 / src GPU 1)
            ...
          ],
          "sum": <float>
        }
      ]
    }
  }

For custom parallel tests, addTestcaseResults is called once per stream, appending
one row each time.  Row order matches the order of addTestcaseResults calls in C++.

Row semantics per custom test category
---------------------------------------
HD_BIDIR   (ce_ce / ce_sm / sm_ce / sm_sm)  : row0=H2D, row1=D2H, row2=Total
H2D        (ce_ce / ce_sm / sm_ce / sm_sm)  : row0=stream0, row1=stream1, row2=Total
D2H        (ce_ce / ce_sm / sm_ce / sm_sm)  : row0=stream0, row1=stream1, row2=Total
D2D_R      (ce_ce / ce_sm / sm_ce / sm_sm)  : row0=stream0, row1=stream1, row2=Total
D2D_W      (ce_ce / ce_sm / sm_ce / sm_sm)  : row0=stream0, row1=stream1, row2=Total
H2D_D2D_R  (*_*)                             : row0=H2D, row1=D2D_Read, row2=Total
H2D_D2D_W  (*_*)                             : row0=H2D, row1=D2D_Write, row2=Total
D2H_D2D_R  (*_*)                             : row0=D2H, row1=D2D_Read, row2=Total
D2H_D2D_W  (*_*)                             : row0=D2H, row1=D2D_Write, row2=Total
CONCURRENT (concurrent_ce)                   : row0..7 = 8 named streams
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not found. Install with:  pip install openpyxl")

# ── constants ─────────────────────────────────────────────────────────────────
RESULTS_DIR = Path("results")

# Row labels for custom parallel tests (3-row pattern: stream0, stream1, total)
# Key = suffix pair extracted from test name
_STREAM_LABELS_2 = {
    ("ce", "ce"):  ["CE stream0", "CE stream1", "Total"],
    ("ce", "sm"):  ["CE stream",  "SM stream",  "Total"],
    ("sm", "ce"):  ["SM stream",  "CE stream",  "Total"],
    ("sm", "sm"):  ["SM stream0", "SM stream1", "Total"],
}
_COMBO_LABELS = {
    "H2D_D2D_R": ["H2D stream", "D2D Read stream",  "Total"],
    "H2D_D2D_W": ["H2D stream", "D2D Write stream", "Total"],
    "D2H_D2D_R": ["D2H stream", "D2D Read stream",  "Total"],
    "D2H_D2D_W": ["D2H stream", "D2D Write stream", "Total"],
    "HD_BIDIR":  ["H2D stream", "D2H stream",        "Total"],
}
_CONCURRENT_LABELS = [
    "S0: CPU→GPU0 (H2D)",
    "S1: GPU0→CPU (D2H)",
    "S2: GPU1→GPU0 read",
    "S3: GPU0→GPU1 write",
    "S4: GPU1→GPU2 write",
    "S5: GPU1→GPU3 write",
    "S6: GPU2→GPU1 read",
    "S7: GPU3→GPU1 read",
]

# ── Excel style helpers ───────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_SUB_FILL   = PatternFill("solid", fgColor="2E5090")
_SUB_FONT   = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL   = PatternFill("solid", fgColor="DCE6F1")
_TOT_FILL   = PatternFill("solid", fgColor="E2EFDA")
_NA_FILL    = PatternFill("solid", fgColor="F2F2F2")
_THIN       = Side(style="thin", color="AAAAAA")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center")

def _hdr(ws, row, col, value, fill=None, font=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = fill  or _HDR_FILL
    cell.font   = font  or _HDR_FONT
    cell.alignment = align or _CENTER
    cell.border = _BORDER
    return cell

def _val(ws, row, col, value, fill=None, is_na=False):
    cell = ws.cell(row=row, column=col, value=value)
    if is_na:
        cell.fill = _NA_FILL
    elif fill:
        cell.fill = fill
    cell.alignment = _CENTER
    cell.border = _BORDER
    if isinstance(value, float):
        cell.number_format = "0.00"
    return cell

def _auto_width(ws, min_w=10, max_w=30):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            max(min_w, min(max_w, length + 2))

# ── JSON parsing ──────────────────────────────────────────────────────────────
def _parse_float(s):
    """Return float or None for 'N/A'."""
    try:
        return float(s)
    except (ValueError, TypeError):
        return None

def load_json(path: Path):
    """Return the first testcase dict from an nvbandwidth JSON file, or None."""
    try:
        with open(path) as f:
            data = json.load(f)
        testcases = data.get("nvbandwidth", {}).get("testcases", [])
        if not testcases:
            return None
        tc = testcases[0]
        if tc.get("status") not in ("Passed",):
            return None
        matrix_raw = tc.get("bandwidth_matrix", [])
        matrix = []
        for row in matrix_raw:
            matrix.append([_parse_float(v) for v in row])
        tc["_matrix"] = matrix
        return tc
    except Exception as e:
        print(f"  [WARN] Failed to parse {path}: {e}")
        return None

# ── row-label inference ───────────────────────────────────────────────────────
def _row_labels(category: str, test_name: str, n_rows: int):
    """Return a list of n_rows human-readable row labels."""
    cat = category.upper()

    # CONCURRENT: always 8 named streams
    if cat == "CONCURRENT" or test_name == "concurrent_ce":
        return _CONCURRENT_LABELS[:n_rows]

    # combo categories: H2D_D2D_R / H2D_D2D_W / D2H_D2D_R / D2H_D2D_W / HD_BIDIR
    if cat in _COMBO_LABELS:
        labels = _COMBO_LABELS[cat]
        return (labels + [f"Row {i}" for i in range(len(labels), n_rows)])[:n_rows]

    # 2-stream H2D / D2H / D2D categories: infer stream types from test suffix
    # e.g. "host_to_device_ce_sm" -> suffix pair ("ce","sm")
    parts = test_name.split("_")
    if len(parts) >= 2:
        pair = (parts[-2], parts[-1])
        if pair in _STREAM_LABELS_2:
            labels = _STREAM_LABELS_2[pair]
            return (labels + [f"Row {i}" for i in range(len(labels), n_rows)])[:n_rows]

    # fallback
    return [f"Row {i}" for i in range(n_rows)]

# ── standard-test sheet builder ───────────────────────────────────────────────
def build_standard_sheet(ws, records):
    """
    records: list of (test_name, tc_dict)
    Standard tests have a bandwidth_matrix of shape (src_gpus × dst_gpus).
    One row per (test_name, src_GPU) pair.
    """
    if not records:
        return

    # Determine max columns across all matrices
    max_cols = max(
        len(row)
        for _, tc in records
        for row in tc["_matrix"]
    ) if records else 0

    # Header row
    _hdr(ws, 1, 1, "Test Name")
    _hdr(ws, 1, 2, "Src / Row")
    for c in range(max_cols):
        _hdr(ws, 1, 3 + c, f"GPU {c}")

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    cur_row = 2
    for i, (test_name, tc) in enumerate(records):
        matrix = tc["_matrix"]
        fill = _ALT_FILL if i % 2 == 0 else None
        desc = tc.get("bandwidth_description", "")

        for r_idx, row_data in enumerate(matrix):
            row_label = f"GPU {r_idx}" if len(matrix) > 1 else "—"
            _val(ws, cur_row, 1, test_name, fill=fill)
            _val(ws, cur_row, 2, row_label, fill=fill)
            ws.cell(row=cur_row, column=1).alignment = _LEFT
            for c_idx, v in enumerate(row_data):
                is_na = v is None
                _val(ws, cur_row, 3 + c_idx,
                     v if v is not None else "N/A",
                     fill=_NA_FILL if is_na else fill,
                     is_na=is_na)
            cur_row += 1

        # Blank separator row between tests
        cur_row += 1

    _auto_width(ws)


# ── custom-test sheet builder ─────────────────────────────────────────────────
def build_custom_sheet(ws, category, records):
    """
    records: list of (test_name, tc_dict)
    Custom tests have bandwidth_matrix rows = one row per parallel stream.
    """
    if not records:
        return

    max_cols = max(
        len(row)
        for _, tc in records
        for row in tc["_matrix"]
    ) if records else 0

    # Header
    _hdr(ws, 1, 1, "Test Name")
    _hdr(ws, 1, 2, "Stream")
    for c in range(max_cols):
        _hdr(ws, 1, 3 + c, f"GPU {c}")

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    cur_row = 2
    for i, (test_name, tc) in enumerate(records):
        matrix = tc["_matrix"]
        labels = _row_labels(category, test_name, len(matrix))
        fill = _ALT_FILL if i % 2 == 0 else None

        for r_idx, (row_data, label) in enumerate(zip(matrix, labels)):
            is_total = "total" in label.lower() or label.lower().startswith("s7")
            row_fill = _TOT_FILL if is_total else fill

            _val(ws, cur_row, 1, test_name, fill=row_fill)
            ws.cell(row=cur_row, column=1).alignment = _LEFT
            _val(ws, cur_row, 2, label, fill=row_fill)
            ws.cell(row=cur_row, column=2).alignment = _LEFT

            for c_idx, v in enumerate(row_data):
                is_na = v is None
                _val(ws, cur_row, 3 + c_idx,
                     v if v is not None else "N/A",
                     fill=_NA_FILL if is_na else row_fill,
                     is_na=is_na)
            cur_row += 1

        cur_row += 1  # blank separator

    _auto_width(ws)


# ── per-category Excel writer ─────────────────────────────────────────────────
def process_category(cat_dir: Path, is_custom: bool, summary_wb=None):
    """
    Parse all *.json files in cat_dir, write a {category}.xlsx there,
    and (optionally) add a sheet to summary_wb.
    Returns list of (test_name, tc_dict) for external use.
    """
    category = cat_dir.name
    json_files = sorted(cat_dir.glob("*.json"))
    if not json_files:
        print(f"  [SKIP] {cat_dir} — no JSON files found")
        return []

    records = []
    for jf in json_files:
        tc = load_json(jf)
        if tc is None:
            print(f"  [SKIP] {jf.name}")
            continue
        records.append((jf.stem, tc))

    if not records:
        print(f"  [SKIP] {cat_dir} — no valid results")
        return []

    print(f"  [{category}] {len(records)} tests")

    # ── per-category workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = category[:31]  # Excel sheet name limit

    if is_custom:
        build_custom_sheet(ws, category, records)
    else:
        build_standard_sheet(ws, records)

    out_path = cat_dir / f"{category}.xlsx"
    wb.save(out_path)
    print(f"    → {out_path}")

    # ── add sheet to summary workbook ─────────────────────────────────────────
    if summary_wb is not None:
        prefix = "custom_" if is_custom else "std_"
        sheet_name = (prefix + category)[:31]
        ws_sum = summary_wb.create_sheet(title=sheet_name)
        if is_custom:
            build_custom_sheet(ws_sum, category, records)
        else:
            build_standard_sheet(ws_sum, records)

    return records


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    if not RESULTS_DIR.exists():
        sys.exit(f"ERROR: '{RESULTS_DIR}' directory not found. "
                 "Run run.sh first.")

    summary_wb = openpyxl.Workbook()
    summary_wb.remove(summary_wb.active)   # remove default empty sheet

    total_tests = 0

    for top in ("standard", "custom"):
        top_dir = RESULTS_DIR / top
        if not top_dir.exists():
            print(f"[INFO] {top_dir} not found, skipping")
            continue

        is_custom = (top == "custom")
        print(f"\n── {top.upper()} ────────────────────────────────────────────")

        for cat_dir in sorted(top_dir.iterdir()):
            if not cat_dir.is_dir():
                continue
            records = process_category(cat_dir, is_custom, summary_wb)
            total_tests += len(records)

    # ── summary Excel ─────────────────────────────────────────────────────────
    if len(summary_wb.sheetnames) == 0:
        print("\n[WARN] No data found; summary not written.")
    else:
        summary_path = RESULTS_DIR / "summary.xlsx"
        summary_wb.save(summary_path)
        print(f"\n✓ Summary: {summary_path}  ({total_tests} tests, "
              f"{len(summary_wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
